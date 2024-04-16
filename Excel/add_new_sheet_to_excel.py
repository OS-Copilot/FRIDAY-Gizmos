import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


def add_new_sheet_to_excel(file_path: str, sheet_name: str) -> None:
    """
    Adds a new sheet with the specified name to an Excel workbook.

    Args:
    file_path (str): Path to the Excel file.
    sheet_name (str): Name of the new sheet to be added.

    Returns:
        None
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Check if the sheet name already exists
        if sheet_name in workbook.sheetnames:
            raise ValueError(f"A sheet named '{sheet_name}' already exists in the workbook.")

        # Add the new sheet
        workbook.create_sheet(sheet_name)

        # Save the workbook
        workbook.save(file_path)

    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' was not found.")
    except InvalidFileException:
        raise InvalidFileException(f"The file '{file_path}' is not a valid Excel file.")

