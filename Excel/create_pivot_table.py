import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


def create_pivot_table(excel_path, source_sheet, target_sheet, row_labels, col_labels, aggfunc):
    """
    Creates a simulated pivot table in an Excel file using openpyxl and pandas.

    Args:
    - excel_path: str - Path to the Excel file.
    - source_sheet: str - Name of the sheet to use as data source.
    - target_sheet: str - Name of the sheet where the pivot table will be created.
    - row_labels: list of str - Columns to be used as row labels.
    - col_labels: list of str - Columns to be used as column labels.
    - aggfunc: str or function - Function to summarize data (e.g., 'mean' for average).

    Returns:
    - None: The function directly modifies the Excel file.
    """
    # Load the workbook and the source sheet
    wb = openpyxl.load_workbook(excel_path)
    source = wb[source_sheet]

    # Read data from source sheet into a DataFrame
    data = pd.DataFrame(source.values)
    # Set the first row as the column names and remove the first row from the DataFrame
    header = data.iloc[0]
    data = data[1:]
    data.columns = header

    # Convert columns to appropriate types
    for col in col_labels:
        data[col] = pd.to_numeric(data[col], errors='coerce')  # Coerce any errors in conversion to NaN

    # Create a pivot table in DataFrame format
    pivot_table = pd.pivot_table(
        data,
        index=row_labels,   # Rows for pivot table
        values=col_labels,  # Values to aggregate
        aggfunc=aggfunc     # Aggregation function
    ).reset_index()  # Reset index to turn the row labels into a column

    # Make sure target sheet is clear
    if target_sheet in wb.sheetnames:
        del wb[target_sheet]
    target = wb.create_sheet(target_sheet)

    # Write the pivot table DataFrame to the target sheet
    for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            target.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(excel_path)
